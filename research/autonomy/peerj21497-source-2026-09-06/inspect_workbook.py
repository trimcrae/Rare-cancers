"""Read-only source inventory; no expression contrasts or inferential analysis."""
import pathlib, json, hashlib, zipfile, collections, math, csv, io
from lxml import etree as ET
import openpyxl

ROOT=pathlib.Path(__file__).resolve().parent
PANEL=('CHRNA6','CD276','SSTR2','PRAME','FAP','CD248','CSPG4','MSLN','L1CAM','GPC3','ALPP','CDH17')
SOURCE=ROOT/'peerj-14-21497-s009.xlsx'
NS={'s':'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

def direct_cells(blob):
    """Independent XML decoder preserves original lexical numeric strings."""
    with zipfile.ZipFile(io.BytesIO(blob)) as z:
        ss=ET.fromstring(z.read('xl/sharedStrings.xml')) if 'xl/sharedStrings.xml' in z.namelist() else None
        strings=[''.join(e.itertext()) for e in ss] if ss is not None else []
        wb=ET.fromstring(z.read('xl/workbook.xml'))
        rel=ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
        rels={e.get('Id'):e.get('Target') for e in rel}
        out={}
        for sh in wb.find('s:sheets',NS):
            p=rels[sh.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')]
            p=p.lstrip('/') if p.startswith('/') else 'xl/'+p
            root=ET.fromstring(z.read(p)); cells={}
            for c in root.findall('.//s:sheetData/s:row/s:c',NS):
                t=c.get('t','n'); v=c.find('s:v',NS)
                if t=='s': val=strings[int(v.text)]
                elif t=='inlineStr': val=''.join(c.find('s:is',NS).itertext())
                else: val=v.text if v is not None else None
                cells[c.get('r')]={'type':t,'value':val,'formula':c.find('s:f',NS) is not None}
            out[sh.get('name')]={'state':sh.get('state','visible'),'part':p,'cells':cells}
        return out,z.namelist()

def inspect():
    raw=SOURCE.read_bytes(); direct,members=direct_cells(raw)
    wb=openpyxl.load_workbook(SOURCE,read_only=False,data_only=False)
    result={'file':SOURCE.name,'sha256':hashlib.sha256(raw).hexdigest(),'md5':hashlib.md5(raw).hexdigest(),'bytes':len(raw),'sheets':[],'defined_names':list(wb.defined_names),'external_links':len(wb._external_links),'archive_members':members}
    features=[]; panel={s:[] for s in PANEL}; checks=0
    for s in wb:
        ds=direct[s.title]; types=collections.Counter(); values=[]; labels=[]; formula=[]; errors=[]; comments=[]
        for row in s:
            for c in row:
                if c.value is None: continue
                d=ds['cells'][c.coordinate]; types[c.data_type]+=1
                if d['type']=='n':
                    assert isinstance(c.value,(int,float)) and float(d['value'])==c.value,(c.coordinate,d,c.value)
                else: assert d['value']==c.value,(c.coordinate,d,c.value)
                checks+=1
                if c.data_type=='f': formula.append(c.coordinate)
                if c.data_type=='e': errors.append(c.coordinate)
                if c.comment: comments.append(c.coordinate)
        for row in s.iter_rows(min_row=2):
            c=row[0]; labels.append(c.value)
            features.append({'sheet':s.title,'cell':c.coordinate,'symbol':c.value,'cell_type':c.data_type,'number_format':c.number_format})
            values.extend([x.value for x in row[1:]])
            if c.value in panel:
                panel[c.value].append({'sheet':s.title,'symbol_cell':c.coordinate,'source_range':f'B{c.row}:M{c.row}','samples':[{'sample':s.cell(1,x.column).value,'cell':x.coordinate,'source_numeric_string':ds['cells'][x.coordinate]['value']} for x in row[1:]]})
        counts=collections.Counter(labels)
        result['sheets'].append({'sheet':s.title,'state':ds['state'],'dimension':s.calculate_dimension(),'rows_including_header':s.max_row,'columns':s.max_column,'header':[{'cell':c.coordinate,'value':c.value} for c in s[1]],'gene_rows':len(labels),'unique_symbols':len(counts),'duplicate_symbols':{k:v for k,v in counts.items() if v>1},'empty_symbol_rows':labels.count(None),'cell_types':dict(types),'formulas':formula,'errors':errors,'comments':comments,'merged_ranges':[str(r) for r in s.merged_cells.ranges],'numeric_values':sum(isinstance(v,(int,float)) for v in values),'missing_values':values.count(None),'nonfinite_values':sum(isinstance(v,(int,float)) and not math.isfinite(v) for v in values),'noninteger_values':sum(isinstance(v,(int,float)) and v!=int(v) for v in values),'negative_values':sum(isinstance(v,(int,float)) and v<0 for v in values)})
    result['panel']=[{'target':s,'status':'reported_processed_row_present' if panel[s] else 'not_in_released_matrix_reason_unknown','rows':panel[s]} for s in PANEL]
    result['nonstring_feature_labels']=[f for f in features if not isinstance(f['symbol'],str)]
    result['string_feature_labels']=sum(isinstance(f['symbol'],str) for f in features)
    result['independent_decoder_checks']={'all_nonempty_cells_compared':checks,'result':'passed'}
    with zipfile.ZipFile(ROOT/'supplementaryFiles.response') as z:
        assert z.read(SOURCE.name)==raw
    article=ET.parse(ROOT/'article.xml'); supp=article.xpath('//*[@id="supp-9"]')[0]
    pi={p.target:p.text for p in supp.xpath('.//processing-instruction()')}
    assert int(pi['suppdata-size'])==len(raw)
    assert pi['suppdata-md5']==hashlib.md5(raw).hexdigest()
    result['article_supp9_integrity']={'size_matches':True,'md5_matches':True,'locator':'//*[@id="supp-9"]/media'}
    terms=('upper quartile','low-expressed','MI-ONCOSEQ','Siriraj Hospital','9,909','22,537','Raw sequencing data','correcting for batch effects','22 patients','12 EMC patients')
    locators=[{'xpath':article.getpath(p),'text':''.join(p.itertext())} for p in article.xpath('//p') if any(t.lower() in ''.join(p.itertext()).lower() for t in terms)]
    (ROOT/'article-source-locators.json').write_text(json.dumps(locators,indent=2,ensure_ascii=False)+'\n',encoding='utf-8')
    result['article_supplement_captions']=[{'id':s.get('id'),'caption':''.join(s.find('caption').itertext())} for s in article.xpath('//supplementary-material')]
    result['article_sample_label_text_hits']={v:article.xpath('string(.)').count(v) for v in [c.value for c in wb.worksheets[0][1]][1:]}
    for name,data in [('workbook-inventory.json',result),('feature-cell-inventory.json',features)]:
        (ROOT/name).write_text(json.dumps(data,indent=2,ensure_ascii=False)+'\n',encoding='utf-8')
    print(json.dumps({'sheets':result['sheets'],'panel':[(s,len(panel[s])) for s in PANEL],'validation':result['independent_decoder_checks']},indent=2))

if __name__=='__main__': inspect()
