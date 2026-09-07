"""Offline structural and byte checks. No expression values are computed."""
from pathlib import Path
import hashlib,json,zipfile,xml.etree.ElementTree as ET
from pypdf import PdfReader
import openpyxl
p=Path(__file__).parent
checks=[]
receipt=json.loads((p/'article-retrieval.json').read_text(encoding='utf-8-sig'))
assert hashlib.sha256((p/'article.xml').read_bytes()).hexdigest()==receipt['sha256']
ET.parse(p/'article.xml');checks.append('Article XML parses and matches original coordinator receipt SHA256')
for rec in json.loads((p/'retrieval-log.json').read_text()):
 if 'sha256' in rec:
  b=(p/rec['file']).read_bytes();assert hashlib.sha256(b).hexdigest()==rec['sha256'];assert len(b)==rec['bytes']
checks.append('All successful retrieval/member hashes and byte counts match')
with zipfile.ZipFile(p/'supplementary-response.bin') as z:
 assert z.testzip() is None
 for name in ['CAC2-45-1760-s001.pdf','CAC2-45-1760-s002.xlsx']:assert z.read(name)==(p/name).read_bytes()
checks.append('ZIP CRC passes and extracted PDF/XLSX equal exact archived bytes')
assert (p/'CAC2-45-1760-s001.pdf').read_bytes().startswith(b'%PDF-');assert len(PdfReader(p/'CAC2-45-1760-s001.pdf').pages)==39
checks.append('PDF parses as 39 pages; methods pages3/4/12/13 rendered and visually inspected separately')
w=openpyxl.load_workbook(p/'CAC2-45-1760-s002.xlsx',read_only=True,data_only=False)
a=json.loads((p/'xlsx-inventory.json').read_text())
assert len(w.sheetnames)==10
for s,v in zip(w,a):
 assert s.title==v['sheet'] and s.max_row==v['rows'] and s.max_column==v['columns']
 assert list(s.iter_rows(max_row=len(v['header_rows']),values_only=True))==[tuple(r) for r in v['header_rows']]
checks.append('XLSX ten sheet dimensions and header-only inventory match')
assert [r['name'] for r in json.loads((p/'github-contents.json').read_text())]==['LICENSE','Main','README.md']
assert len(json.loads((p/'github-main.json').read_text()))==5
checks.append('Frozen GitHub root and Main lists contain3 and5 entries')
assert 'EGAD50000001419' in (p/'ega-dac-response.html').read_text()
checks.append('Exact EGA public DAC response contains dataset EGAD50000001419')
print(json.dumps({'status':'pass','checks':checks},indent=2))
