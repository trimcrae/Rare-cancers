from pathlib import Path
import io,json,zipfile,xml.etree.ElementTree as ET,sys
from openpyxl import load_workbook
sys.stdout.reconfigure(encoding='utf-8')
root=Path.cwd()
r=ET.parse(root/'research/autonomy/atlas-hofvander-source-2026-09-06/hofvander2026.xml').getroot()
for s in r.findall('.//sec'):
 title=s.find('title')
 if title is not None and any(x in ''.join(title.itertext()).lower() for x in ['material','patient','rna sequenc','data availability']):print('SOURCE SECTION',' '.join(s.itertext()))
with zipfile.ZipFile(root/'research/autonomy/atlas-hofvander-source-2026-09-06/source-provenance.zip') as z:
 wb=load_workbook(io.BytesIO(z.read('ccr-25-3740_supplementary_table_s1_suppts1.xlsx')),read_only=True,data_only=True)
 print('SHEETS',wb.sheetnames)
 for row in wb.worksheets[0].iter_rows():
  vals=[c.value for c in row]
  if row[0].row<6 or any('myxoid chondrosarcoma' in str(x).lower() for x in vals):print('ROW',row[0].row,vals)
 print('META_HEAD',z.read('source_data/meta_data.txt').decode().splitlines()[:4])
wb=load_workbook(root/'research/autonomy/atlas-prior-art-2026-09-06/428_2023_3606_MOESM1_ESM.xlsx',read_only=True,data_only=True)
for i in [1,2,128,129,130,131,132]:print('PRAME',i,[c.value for c in wb.worksheets[0][i]])
with zipfile.ZipFile(root/'research/autonomy/atlas-normal-context-2026-09-06/original-source-packet.zip') as z:
 c=json.loads(z.read('CSPG4.json'));print('HPA_CSPG4',json.dumps(c,ensure_ascii=False))
