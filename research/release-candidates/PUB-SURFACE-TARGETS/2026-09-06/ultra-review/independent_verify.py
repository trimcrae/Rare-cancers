from pathlib import Path
import bisect,collections,csv,gzip,hashlib,io,json,math,sys,zipfile
import numpy as np
from openpyxl import load_workbook
sys.stdout.reconfigure(encoding='utf-8')
ROOT=Path.cwd(); OUT=ROOT/'review-results'; P=ROOT/'research/autonomy/atlas-hofvander-validation-2026-09-06'
read_files={}; members={}; errors=[]; checks=collections.Counter()
def read(p):
 p=Path(p); b=p.read_bytes(); read_files[str(p.relative_to(ROOT)).replace('\\','/')]=hashlib.sha256(b).hexdigest();return b
def j(p):return json.loads(read(p))
def member(p,n):
 with zipfile.ZipFile(io.BytesIO(read(p))) as z:b=z.read(n)
 members[str(p.relative_to(ROOT)).replace('\\','/')+'!'+n]=hashlib.sha256(b).hexdigest();return b
def equal(label,a,b,tol=1e-12):
 checks[label.split('/')[0]]+=1
 ok=(a==b) if a is None or b is None or not isinstance(a,(int,float)) or isinstance(a,bool) else abs(a-b)<=tol
 if not ok:errors.append({'field':label,'independent':a,'reported':b})
EMC='Extraskeletal myxoid chondrosarcoma'; HS=['Myxoid liposarcoma','Low-grade fibromyxoid sarcoma','Synovial sarcoma']; SECOND=['Myxofibrosarcoma','Solitary fibrous tumor','Desmoid']; CONTEXT=['Myxofibrosarcoma','Dermatofibrosarcoma protuberans']
GENES='CD276 SSTR2 PRAME FAP CD248 CSPG4 MSLN L1CAM GPC3 ALPP CDH17 CHRNA6'.split(); OLD={'104-92','168-97','536-00'}
hzip=ROOT/'research/autonomy/atlas-hofvander-source-2026-09-06/source-provenance.zip'
metadata=list(csv.DictReader(io.StringIO(member(hzip,'source_data/meta_data.txt').decode()),delimiter='\t'))
byid={r['lab_no']:r for r in metadata}
wb=load_workbook(io.BytesIO(member(hzip,'ccr-25-3740_supplementary_table_s1_suppts1.xlsx')),read_only=True,data_only=True)
rows=[]
for index,r in enumerate(wb.active.iter_rows(values_only=True),1):
 sid=str(r[0]).split('_')[0]
 if sid in byid:rows.append({'sample_id':sid,'diagnosis':r[1] or '', 'revised_diagnosis':r[2] or '', 'sequencing_year':byid[sid]['sequencing_year'],'specimen_exception':r[12] or '', 'prior_comment':r[21] or '', 'eligible':not r[12] and sid not in OLD,'s1_row':index})
hm=j(P/'metadata-manifest.json')
equal('metadata/count',len(rows),704)
for a,b in zip(rows,hm['samples']):
 for k in a:equal('metadata/'+a['sample_id']+'/'+k,a[k],b[k])
retained=[r for r in rows if r['eligible']]
raw=gzip.decompress(read(ROOT/'research/autonomy/atlas-hofvander-source-2026-09-06/tpm_matrix.tsv.gz'))
equal('source/tpm_uncompressed_hash',hashlib.sha256(raw).hexdigest(),hm['source_files']['source_data/tpm_matrix.tsv']['sha256'])
reader=csv.reader(io.StringIO(raw.decode()),delimiter='\t');header=next(reader)[1:];val={};rowcount=0
for r in reader:
 rowcount+=1
 if r[0] in GENES:
  assert r[0] not in val and len(r)==705
  val[r[0]]=dict(zip(header,map(float,r[1:])))
equal('source/feature_count',rowcount,19116);equal('source/sample_count',len(header),704)
expected=j(P/'results/selected-values.json')
for g in GENES:
 for sid,x in val[g].items():equal('TPM_values/'+g+'/'+sid,x,expected[g][sid])
def auc(x,y):
 if not x or not y:return None
 sy=sorted(y);u=sum(bisect.bisect_left(sy,a)+.5*(bisect.bisect_right(sy,a)-bisect.bisect_left(sy,a)) for a in x)
 return u/(len(x)*len(y))
def calc(rr,v,h):
 ee=[r for r in rr if r['diagnosis']==EMC];cc=[r for r in rr if r['diagnosis']==h]
 cells=[]
 for yr in sorted(set(r['sequencing_year'] for r in ee+cc)):
  e=[r for r in ee if r['sequencing_year']==yr];c=[r for r in cc if r['sequencing_year']==yr]
  a=auc([v[r['sample_id']] for r in e],[v[r['sample_id']] for r in c]);cells.append((yr,len(e),len(c),a))
 denom=sum(e for yr,e,c,a in cells if a is not None)
 return {'marginal':auc([v[r['sample_id']] for r in ee],[v[r['sample_id']] for r in cc]),'matched':sum(e*a for yr,e,c,a in cells if a is not None)/denom if denom else None},cells

def summarize(rr,v,hs=HS):
 hv={h:calc(rr,v,h)[0] for h in hs}
 return {m:sum(hv[h][m] for h in hs)/len(hs) if all(hv[h][m] is not None for h in hs) else None for m in ['marginal','matched']},hv
passes=[];computed={}
for g in GENES:
 result=j(P/'results'/f'{g}.json');v=val[g];summary,hv=summarize(retained,v)
 for m in summary:equal('primary_summary/'+g+'/'+m,summary[m],result['primary']['summary'][m])
 for key,hs in [('primary',HS),('context',CONTEXT)]:
  for h in hs:
   av,cells=calc(retained,v,h);ex=result[key]['histologies'][h]
   for m in av:equal('histology/'+g+'/'+h+'/'+m,av[m],ex[m])
   for cell,ec in zip(cells,ex['cells']):
    for k,a in zip(['year','n_emc','n_comparator','A'],cell):equal('yearcell/'+g+'/'+h+'/'+str(cell[0])+'/'+k,a,ec[k])
 deletionvals={}
 for kind,ds in result['deletions'].items():
  deletionvals[kind]=[]
  for entry in ds:
   d=entry['deleted'];rr=[r for r in retained if r['sequencing_year']!=d] if kind=='year' else [r for r in retained if r['sample_id']!=d]
   hs=[h for h in HS if h!=d] if kind=='histology' else HS
   sm,hvs=summarize(rr,v,hs);deletionvals[kind].append(sm)
   for m in sm:equal('deletion_summary/'+g+'/'+kind+'/'+d+'/'+m,sm[m],entry['summary'][m])
   for h,vs in hvs.items():
    for m,a in vs.items():equal('deletion_histology/'+g+'/'+kind+'/'+d+'/'+h+'/'+m,a,entry['histologies'][h][m])
 for kind,entries in deletionvals.items():
  for m in ['marginal','matched']:
   defined=[a[m] for a in entries if a[m] is not None]
   for n,a in enumerate([min(defined),max(defined)]):equal('deletion_range/'+g+'/'+kind+'/'+m+str(n),a,result['deletion_summary_ranges'][kind][m][n])
 if g!='CHRNA6':
  passed=summary['marginal']>=.7 and all(a[m] is not None and a[m]>.5 for a in hv.values() for m in ['marginal','matched']) and all(s[m] is not None and s[m]>.5 for k in ['emc','histology'] for s in deletionvals[k] for m in ['marginal','matched'])
  equal('allocation/'+g,passed,result['allocation_rule']['consistent_RNA_rationale'])
  if passed:passes.append(g)
 revised=[dict(r,diagnosis=r['revised_diagnosis'] or r['diagnosis']) for r in retained]
 sm,_=summarize(revised,v)
 for m in sm:equal('revision/'+g+'/'+m,sm[m],result['revised_diagnosis_sensitivity']['summary'][m])
 computed[g]={'primary':summary,'LGFMS':hv[HS[1]]}
# Original SOFT parser: retain sample headers and original values, never the analysis cache.
am=j(P/'replication-manifest.json');probes=am['gene_to_probe'];byprobe={p:g for g,p in probes.items()}
soft=gzip.decompress(read(ROOT/'research/autonomy/atlas-primary-provenance-2026-09-06/GSE24369.soft.gz')).decode()
array={g:{} for g in GENES};array_labels={}
for block in soft.split('^SAMPLE = ')[1:]:
 lines=block.splitlines();sid=lines[0].strip();title=next(x.split(' = ',1)[1] for x in lines if x.startswith('!Sample_title = '));array_labels[sid]=title
 st=lines.index('!sample_table_begin');end=lines.index('!sample_table_end');cols=lines[st+1].split('\t')
 for line in lines[st+2:end]:
  fs=line.split('\t');probe=fs[cols.index('ID_REF')]
  if probe in byprobe:array[byprobe[probe]][sid]=float(fs[cols.index('VALUE')])
ax=j(P/'replication-results/array-values.json')
for g in GENES:
 for sid,v in array[g].items():equal('array_values/'+g+'/'+sid,v,ax[g][sid])
# Recover diagnosis independently from the sample-title wording.
def diag(title):
 t=title.lower()
 if 'extraskeletal myxoid chondrosarcoma' in t:return EMC
 if 'low grade fibromyxoid sarcoma' in t or 'low-grade fibromyxoid sarcoma' in t:return HS[1]
 if 'myxofibrosarcoma' in t:return SECOND[0]
 if 'solitary fibrous tumor' in t:return SECOND[1]
 if 'desmoid' in t:return SECOND[2]
 if 'muscle' in t:return 'Normal muscle pool'
 raise ValueError(title)
aro=[{'sample_id':sid,'diagnosis':diag(t)} for sid,t in array_labels.items()]
for g in GENES:
 ex=j(P/'replication-results'/f'{g}.json')
 for h,contrast in ex['contrasts'].items():
  e=[r['sample_id'] for r in aro if r['diagnosis']==EMC];c=[r['sample_id'] for r in aro if r['diagnosis']==h]
  equal('array_contrast/'+g+'/'+h,auc([array[g][s] for s in e],[array[g][s] for s in c]),contrast['array']['A'])
  for kind,ids in [('EMC',e),('comparator',c)]:
   for sid,dd in zip(ids,contrast['array']['deletions'][kind]):equal('array_deletion/'+g+'/'+h+'/'+sid,auc([array[g][s] for s in e if s!=sid],[array[g][s] for s in c if s!=sid]),dd['A'])
  hv,_=calc(retained,val[g],h)
  for m,a in hv.items():equal('shared_hofvander/'+g+'/'+h+'/'+m,a,contrast['hofvander']['summary'][m])
# Independently resample stratum counts, sharing EMC weights across comparisons.
B=10000;rng=np.random.default_rng(2026090601);rr=[r for r in retained if r['diagnosis'] in [EMC]+HS];w=np.zeros((B,len(rr)))
for key in sorted(set((r['diagnosis'],r['sequencing_year']) for r in rr)):
 idx=[i for i,r in enumerate(rr) if (r['diagnosis'],r['sequencing_year'])==key]
 w[:,idx]=rng.multinomial(len(idx),[1/len(idx)]*len(idx),size=B)
be=[];ce=[];eid=[i for i,r in enumerate(rr) if r['diagnosis']==EMC]
for h in HS:
 cid=[i for i,r in enumerate(rr) if r['diagnosis']==h]
 def weighted(e,c):
  x=np.array([val['CSPG4'][rr[i]['sample_id']] for i in e]);y=np.array([val['CSPG4'][rr[i]['sample_id']] for i in c]);R=(x[:,None]>y[None,:])+.5*(x[:,None]==y[None,:])
  return np.einsum('bi,ij,bj->b',w[:,e],R,w[:,c])/(len(e)*len(c))
 be.append(weighted(eid,cid));matched=np.zeros(B);den=0
 for yr in sorted(set(r['sequencing_year'] for r in rr)):
  e=[i for i in eid if rr[i]['sequencing_year']==yr];c=[i for i in cid if rr[i]['sequencing_year']==yr]
  if e and c:matched+=len(e)*weighted(e,c);den+=len(e)
 ce.append(matched/den)
bootstrap={'draws':B,'seed':2026090601,'method':'independent multinomial stratum-count resampling shared across histologies','marginal_percentile_95':np.percentile(np.mean(be,axis=0),[2.5,97.5]).tolist(),'matched_percentile_95':np.percentile(np.mean(ce,axis=0),[2.5,97.5]).tolist()}
# Annotation mapping read from original complete platform file.
azip=ROOT/'research/autonomy/atlas-original-array-source-2026-09-06/original-source-recovery.zip'
annotation=member(azip,'GPL6244-original-annotation.tsv').decode();hit=[l for l in annotation.splitlines() if l.split('\t')[0] in byprobe]
(OUT/'selected-original-platform-rows.tsv').write_text('\n'.join(hit)+'\n',encoding='utf-8')
equal('annotation/selected_rows',len(hit),12)
result={'status':'passed' if not errors else 'failed','checks':dict(checks),'total_comparisons':sum(checks.values()),'mismatches':errors,'passing_address_genes':passes,'computed':computed,'independent_bootstrap':bootstrap,'cohort_counts':dict(collections.Counter(r['diagnosis'] for r in retained if r['diagnosis'] in [EMC]+HS+CONTEXT+SECOND)),'array_counts':dict(collections.Counter(r['diagnosis'] for r in aro)),'read_files':read_files,'archive_members':members}
(OUT/'independent-verification.json').write_text(json.dumps(result,indent=2)+'\n',encoding='utf-8');print(json.dumps({k:v for k,v in result.items() if k not in ['read_files','archive_members','computed']},indent=2))
